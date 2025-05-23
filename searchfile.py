# -*- coding: utf-8 -*-
import os
from docx import Document
from PyPDF2 import PdfReader
from openpyxl import load_workbook

def search_docx(file_path, keyword):
    """docx�t�@�C�����ŃL�[���[�h������"""
    try:
        doc = Document(file_path)
        for paragraph in doc.paragraphs:
            if keyword in paragraph.text:
                return True
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
    return False

def search_pdf(file_path, keyword):
    """PDF�t�@�C�����ŃL�[���[�h������"""
    try:
        reader = PdfReader(file_path)
        for page in reader.pages:
            if keyword in page.extract_text():
                return True
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
    return False

def search_xlsx(file_path, keyword):
    """xlsx�t�@�C�����ŃL�[���[�h������"""
    try:
        wb = load_workbook(file_path, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and keyword in str(cell):
                        return True
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
    return False

def find_files_with_keyword(directory, keyword):
    """�f�B���N�g�����̎w�肵���t�@�C���`������L�[���[�h������"""
    result = []
    for root, _, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            if file.endswith(".docx"):
                if search_docx(file_path, keyword):
                    result.append(file)
            elif file.endswith(".pdf"):
                if search_pdf(file_path, keyword):
                    result.append(file)
            elif file.endswith(".xlsx"):
                if search_xlsx(file_path, keyword):
                    result.append(file)
    return result

# �����Ώۂ̃f�B���N�g���ƃL�[���[�h���w��
directory = "C:\\path\\to\\your\\folder"
keyword = "�����������L�[���[�h"

# ���ʂ��o��
files = find_files_with_keyword(directory, keyword)
print("Found files:", files)
